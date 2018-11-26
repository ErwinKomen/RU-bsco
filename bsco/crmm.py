# ==========================================================================================================
# Name :    crmm
# Goal :    Read corpus CRM-Meertens information from an Excel
#           1 - convert this information into a JSON file
#           2 - act on this information by 
#               a. reading the correct files from the internet
#               b. saving them correctly
# History:
# 26/nov/2018    ERK Created
# ==========================================================================================================
import sys, getopt, os.path, importlib
import util, json, io
import openpyxl
import requests
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook, load_workbook
from models import CrmmInfo

# ============================= LOCAL VARIABLES ====================================
errHandle = util.ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 26/nov/2018    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input Excel file name
  flOutput = ''       # output directory where .psd and .meta.xml files should come

  try:
    # Adapt the program name to exclude the directory
    index = prgName.rfind("\\")
    if (index > 0) :
      prgName = prgName[index+1:]
    sSyntax = prgName + ' -i <input Excel file> -o <output directory>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hi:o:", ["-inputfile=","-outputdir="])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile", "--inputfile"):
        flInput = arg
      elif opt in ("-o", "--odir", "--outputdir"):
        flOutput = arg
    # Check if all arguments are there
    if (flInput == '' or flOutput == ''):
      errHandle.DoError(sSyntax)
    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')
    # Call the function that does the job
    oArgs = {'input': flInput,
             'output': flOutput}
    if (process_crmm(oArgs)) :
      errHandle.Status("Ready")
    else :
      errHandle.DoError("Could not complete")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    process_crmm
# Goal :    Read the indicated Excel file, convert it to JSON and act upon it
# History:
# 26/nov/2018    ERK Created
# ----------------------------------------------------------------------------------
def process_crmm(oArgs):
    bDoAsk = False  # Local variable
    flInput = ""    # 
    flOutput = ""   # Output JSON file
    dirOutput = ""  # Output directory
    arInput = []    # Array of input files
    arOutput = []   # Array of output files
    lOutput = []    # List of output objects (one per hit)

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: dirOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Double check the output: should be a DIRECTORY
        if os.path.exists(dirOutput):
            # Check if it accidentily is a directory
            if not os.path.isdir(dirOutput):
                errHandle.Status("Please specify an output DIRECTORY")
                return False

        # Create an output file name
        flOutput = dirOutput.split(sep=".")[0]
        flOutput = "{}_info.json".format(flOutput)

        # Open the input file and read it as Excel
        wbInput = openpyxl.load_workbook(filename=flInput)
        # Assume the active worksheet
        ws = wbInput.active
        # Walk through all rows in the worksheet
        row = 2
        bFinish = False
        info_list = []
        info_out = []
        while not bFinish:
            # Get the record number
            recnum = ws.cell(column=1, row=row).value
            bFinish = (recnum == None or recnum == "")
            if recnum == 96 or recnum == "96":
                iStop = 1
            if not bFinish:
                # We can still read this one
                filenum = ws.cell(column=2, row=row).value
                cell_mrg = ws.cell(column=3, row=row)
                cell_meta = ws.cell(column=4, row=row)
                loc_wijk = ws.cell(column=5, row=row).value
                loc_gron = ws.cell(column=6, row=row).value
                loc_have = ws.cell(column=7, row=row).value
                location = get_location(loc_wijk, loc_gron, loc_have)
                # Get the correct file name and link from [cell_mrg]
                mrg_name = cell_mrg.value
                mrg_url = cell_mrg.hyperlink.target
                # Get the correct hyperlink from [cell_meta]
                meta_url = cell_meta.hyperlink.target
                # Create an Information object that holds all of this
                oInfo = CrmmInfo(line=row, filenum=filenum, mrg_name=mrg_name, mrg_url=mrg_url, meta_url=meta_url, location=location)
                # Add this object to a list
                info_list.append(oInfo)
                info_out.append(oInfo.get_json())
            # Go to the next row
            row += 1

        # Write the list of Crmm Info to a JSON file
        with io.open(flOutput, "w", encoding='utf-8-sig') as f:
            json.dump(info_out, f)

        # Walk the list of Crmm Info
        for oInfo in info_list:
            # Download the MRG to the target
            oResult = oInfo.create_psd(dirOutput)
            if oResult['status'] != "ok":
                errHandle.Status("Could not read PSD for {}".format(oInfo['filenum']))
            # Download and save the metadata
            oResult = oInfo.create_meta(dirOutput)
            if oResult['status'] != "ok":
                errHandle.Status("Could not read metadata for {}".format(oInfo['filenum']))

        # We are happy: return okay
        return True
    except:
        # act
        smsg = errHandle.get_error_message()
        errHandle.DoError("process_crmm")
        return False


def get_location(wijk, gron, have):
    location = ""
    if wijk == 1 or wijk == "1":
        location = "xDF575regioDeWijk"
    elif gron == 1 or gron == "1":
        location = "xDC108Groningen1"
    elif have == "" or have == "1":
        location = "xDF573regioHavelte"
    return location

# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
