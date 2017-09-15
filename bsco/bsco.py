# ==========================================================================================================
# Name :    kamer
# Goal :    Extract intensifiers from political debate texts
# History:
# 16/feb/2017    ERK Created
# ==========================================================================================================
import sys, getopt, os.path, importlib
import util, json, io
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook

# ============================= LOCAL VARIABLES ====================================
errHandle = util.ErrHandle()
lStructure = [
        {"name": "Documents", "header": []},
        {"name": "Bookseller", "header": ["Name"]},
        {"name": "Catalogue", "header": ["Annotation", "Pages (original)", "Scanned", "Type",
                                     "Literature", "Pages", "Collation", "Manuscript notes",
                                     "Scanned copy", "Location", "Title", "Location (original)",
                                     "Title page transcription"]},
        {"name": "Citation", "header": ["Title", "URL", "ID", "Type"]},
        {"name": "Owner", "header": ["Profession", "Location", "Name"]},
        {"name": "Auction", "header": ["Venue", "Year", "Location", "Date"]}
        ]

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 15/sep/2017    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input directory name
  flOutput = ''       # output directory name
  sMethod = 'compact' # Output method: "compact", "full"
                      #   compact = all on one worksheet
                      #   full    = use separate worksheets

  try:
    # Adapt the program name to exclude the directory
    index = prgName.rfind("\\")
    if (index > 0) :
      prgName = prgName[index+1:]
    sSyntax = prgName + ' [-m <method>] -i <input directory> -o <output directory>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hm:i:o:", ["-method=","-inputdir=","-outputdir="])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-m", "--method"):
        sMethod = arg
      elif opt in ("-i", "--ifile"):
        flInput = arg
      elif opt in ("-o", "--ofile"):
        flOutput = arg
    # Check if all arguments are there
    if (flInput == '' or flOutput == ''):
      errHandle.DoError(sSyntax)
    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')
    errHandle.Status('Output method is "' + sMethod + '"')
    # Call the function that does the job
    oArgs = {'input': flInput,
             'output': flOutput,
             'method': sMethod}
    if (process_bsco(oArgs)) :
      errHandle.Status("Ready")
    else :
      errHandle.DoError("Could not complete")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    process_bsco
# Goal :    Read the indicated JSON file and convert it to Excel
# History:
# 14/sep/2017    ERK Created
# ----------------------------------------------------------------------------------
def process_bsco(oArgs):
    oAdv = None     # 
    bDoAsk = False  # Local variable
    flInput = ""    # 
    flOutput = ""   # 
    sMethod = ""    # 
    arInput = []    # Array of input files
    arOutput = []   # Array of output files
    lOutput = []    # List of output objects (one per hit)

    try:
        # Recover the arguments
        if "method" in oArgs: sMethod = oArgs["method"]
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Adapt the output file name
        flOutput = flOutput.split(sep=".")[0]
        flOutput = "{}_{}.xlsx".format(flOutput,sMethod)

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Double check the output
        if os.path.exists(flOutput):
            # Check if it accidentily is a directory
            if os.path.isdir(flOutput):
                errHandle.Status("Please specify an output FILE")
                return False
            else:
                # give warning that we will overwrite
                errHandle.Status("We will overwrite the existing output file")

        # Open the input file for reading, and treat it as UTF8 encoded
        f = io.open(flInput, mode="r", encoding="UTF-8")
        # Read the input file as JSON
        oInput = json.load(f)
        # Close the input file again
        f.close()

        # Create an excel file with the correct worksheets
        wbOutput = openpyxl.Workbook()

        # Action depends on the method used
        if sMethod == "full":
            # Divide into separate sheets
            lSheet = []
            bFirst = True
            # Process all the required worksheets
            for sheetobject in lStructure:
                # Get the name of the sheet
                sheetName = sheetobject["name"]
                errHandle.Status("Processing sheet: " + sheetName)

                # Naming depends on whether this is the first sheet
                if bFirst:
                    sheet = wbOutput.get_active_sheet()
                    sheet.title = sheetName
                    bFirst = False
                else:
                    # Create a worksheet
                    sheet = wbOutput.create_sheet(sheetName)

                # Add this sheet to a list
                lSheet.append(sheet)
                # Add headers for this sheet
                lHeader = sheetobject['header']

                # Set the column number: start with 1
                col_num = 1

                # Check if there are columns
                if len(lHeader) == 0:
                    # There are no columns, there is just a list
                    add_header_row(sheet, ["no header"])
                    lInput = oInput[sheetName]
                    add_list(sheet, lInput, col_num)
                else:
                    # There are columns: add their names
                    add_header_row(sheet, lHeader)
                    # Process all the columns
                    for colName in lHeader:
                        errHandle.Status("   column: " + colName)
                        # Get the list of information for this column
                        lInput = oInput[sheetName][colName]
                        # Add this list to the sheet
                        add_list(sheet, lInput, col_num)
                        col_num += 1
        elif sMethod == "compact":
            # Put the output on to one sheet
            sheet = wbOutput.get_active_sheet()
            sheet.title = "Compact"

            # Set the column number: start with 1
            col_num = 1

            # Process all the required worksheets
            for sheetobject in lStructure:
                # Get the name of the sheet
                sheetName = sheetobject["name"]
                errHandle.Status("Processing sheet: " + sheetName)

                # Get the headers for this sheet
                lHeader = sheetobject['header']

                # Check if there is a separate header
                if len(lHeader) == 0:
                    # There are no columns, there is just a list
                    add_one_header(sheet, col_num, sheetName)
                    lInput = oInput[sheetName]
                    add_list(sheet, lInput, col_num)
                    col_num += 1
                else:
                    # Process all the columns
                    for colName in lHeader:
                        errHandle.Status("   column: " + colName)
                        # Add a header for this column
                        add_one_header(sheet, col_num, "{}.{}".format(sheetName, colName ))
                        # Get the list of information for this column
                        lInput = oInput[sheetName][colName]
                        # Add this list to the sheet
                        add_list(sheet, lInput, col_num)
                        col_num += 1

        
        # Save the output file
        errHandle.Status("Saving... "+flOutput)
        wbOutput.save(flOutput)

        # We are happy: return okay
        return True
    except:
        # act
        errHandle.DoError("process_bsco")
        return False
    
def add_header_row(wsThis, lColNames):
    """Add a header row using the names in [lColNames] to worksheet wsThis"""

    try:
        # Iterate through all the intended columns
        for col_num in range(len(lColNames)):
            add_one_header(wsThis, col_num+1, lColNames[col_num])
        # Return okay
        return True
    except:
        # Show the error
        errHandle.DoError("add_header_row")
        return False

def add_one_header(wsThis, col_num, sName):
    """Add a header called [sName] to worksheet wsThis, column [col_num]"""

    try:
        row_num = 1
        # Iterate through all the intended columns
        # Get the correct cell
        c = wsThis.cell(row=row_num, column=col_num)
        # Set the value of this cell correctly
        c.value = sName
        # Set the font of this cell to bold
        c.font = openpyxl.styles.Font(bold=True)
        # Set the width of this column to a standard value
        wsThis.column_dimensions[get_column_letter(col_num)].width = 20.0
        # Return okay
        return True
    except:
        # Show the error
        errHandle.DoError("add_one_header")
        return False

def add_list(wsThis, lThis, col_num):
    """Add the list in [lThis] on worksheet [wsThis] to column [col_num]"""

    try:
        # Walk all items in the list
        row_num = 0
        for item in lThis:
            row_num += 1
            # Get the cell number it needs to be put into
            c = wsThis.cell(row=row_num + 1, column=col_num)
            # Set the value for this cell: this depends on what this is
            if item == None:
                c.value = ""
            elif type(item) is list:
                # If the list is just one member...
                if len(item) == 1:
                    fItem = item[0]
                    if fItem == None:
                        c.value = ""
                    elif type(fItem) is str and fItem.startswith("="):
                        # Items that start with '=' are otherwise treated as formulas
                        c.value = json.dumps(item)
                    else:
                        c.value = fItem
                else:
                    c.value = json.dumps(item)
            elif type(item) is str and item.startswith("="):
                # Items that start with '=' are otherwise treated as formulas
                c.value = json.dumps(item)
            elif type(item) is int or type(item) is str:
                c.value = item
            else:
                c.value = json.dumps(item)
            # Set the alignment
            c.alignment = openpyxl.styles.Alignment(wrap_text=False)

        # Return okay
        return True
    except:
        # Show the error
        errHandle.DoError("add_list")
        return False

# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
